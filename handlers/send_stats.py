import io
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from aiogram import F
from aiogram.enums import ChatAction
from aiogram.types import Message, BufferedInputFile

from loader import db
from router import router

_HEADERS = [
    "ID", "ФИО пользователя", "Телефон пользователя",
    "Тип обращения", "Обращение", "Этаж", "Решён",
    "ФИО решившего", "Телефон решившего", "Дата регистрации",
]

# Column widths in Excel units (base unit * ratio [2,6,6,4,6,2,2,6,6,4])
_COL_WIDTHS = [10, 30, 30, 20, 30, 10, 10, 30, 30, 20]

_CONTENT_TYPE_LABELS = {
    "text":       "Текст",
    "photo":      "Фото",
    "document":   "Файл",
    "voice":      "Голосовое сообщение",
    "video":      "Видео",
    "video_note": "Круглое видео",
}

_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_FONT = Font(bold=True, size=13, color="000000")
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _val(v) -> str:
    if v is None or v == "":
        return "-"
    return v


def _fmt_phone(v) -> str:
    if v is None or v == "":
        return "-"
    d = str(v).strip()
    if len(d) == 9:
        return f"+998 ({d[0:2]}) {d[2:5]}-{d[5:7]}-{d[7:9]}"
    return f"+998 {d}"


@router.message(F.text == "/send_stats")
async def send_stats(message: Message):
    await message.bot.send_chat_action(chat_id=message.chat.id, action=ChatAction.UPLOAD_DOCUMENT)

    incidents = await db.fetchall("""
        SELECT
            i.id,
            u.fullname      AS user_fullname,
            u.phone_number  AS user_phone,
            i.incident_description_type,
            i.incident,
            i.document_caption,
            i.floor,
            i.is_solved,
            s.fullname      AS solver_fullname,
            s.phone_number  AS solver_phone,
            i.datetime
        FROM incidents i
        JOIN  users u ON i.user_id    = u.id
        LEFT JOIN users s ON i.solved_by = s.id
        ORDER BY i.id DESC
    """)

    if not incidents:
        await message.answer(text="Инцидентов пока нет.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по инцидентам"

    for i, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    for row in incidents:
        is_solved = row["is_solved"]
        dt = row["datetime"].strftime("%d.%m.%Y %H:%M") if row["datetime"] else "-"

        ws.append([
            _val(row["id"]),
            _val(row["user_fullname"]),
            _fmt_phone(row["user_phone"]),
            _CONTENT_TYPE_LABELS.get(row["incident_description_type"], _val(row["incident_description_type"])),
            _val(row["document_caption"]) if row["incident_description_type"] in ("photo", "document", "video") else ("-" if row["incident_description_type"] == "video_note" else _val(row["incident"])),
            _val(row["floor"]),
            "Да" if is_solved else "Нет",
            _val(row["solver_fullname"]),
            _fmt_phone(row["solver_phone"]),
            dt,
        ])

        fill = _GREEN_FILL if is_solved else _RED_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = _CENTER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Отчет по инцидентам {datetime.now(ZoneInfo('Asia/Tashkent')).strftime('%d.%m.%Y')}.xlsx"
    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename=filename),
    )
